from decimal import Decimal
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from . import ingresos as mod_ingresos
from . import gastos as mod_gastos
from . import facturas as mod_facturas
from . import informes as mod_informes


_COLOR_CABECERA = "2E4057"
_COLOR_INGRESO = "D4EDDA"
_COLOR_GASTO = "F8D7DA"
_COLOR_TOTAL = "FFF3CD"

_BORDE_FINO = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _cabecera(hoja, fila: int, columnas: list[str]):
    for col_idx, titulo in enumerate(columnas, 1):
        celda = hoja.cell(row=fila, column=col_idx, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=_COLOR_CABECERA)
        celda.alignment = Alignment(horizontal="center")
        celda.border = _BORDE_FINO


def _fila_datos(hoja, fila: int, datos: list, color: str = None):
    for col_idx, valor in enumerate(datos, 1):
        celda = hoja.cell(row=fila, column=col_idx, value=valor)
        celda.border = _BORDE_FINO
        if color:
            celda.fill = PatternFill("solid", fgColor=color)


def _ajustar_columnas(hoja):
    for col in hoja.columns:
        max_ancho = max((len(str(c.value or "")) for c in col), default=10)
        hoja.column_dimensions[get_column_letter(col[0].column)].width = min(max_ancho + 4, 40)


def _hoja_ingresos(libro, anio: int):
    hoja = libro.create_sheet("Ingresos")
    filas = mod_ingresos.listar_ingresos(anio=anio, limite=5000)
    _cabecera(hoja, 1, ["ID", "Fecha", "Importe (€)", "Categoría", "Descripción"])
    for i, fila in enumerate(filas, 2):
        _fila_datos(hoja, i, [
            fila["id"],
            str(fila["fecha"]),
            float(fila["importe"]),
            fila.get("categoria") or "",
            fila.get("descripcion") or "",
        ], _COLOR_INGRESO)
    _ajustar_columnas(hoja)


def _hoja_gastos(libro, anio: int):
    hoja = libro.create_sheet("Gastos")
    filas = mod_gastos.listar_gastos(anio=anio, limite=5000)
    _cabecera(hoja, 1, ["ID", "Fecha", "Importe (€)", "Categoría", "Descripción"])
    for i, fila in enumerate(filas, 2):
        _fila_datos(hoja, i, [
            fila["id"],
            str(fila["fecha"]),
            float(fila["importe"]),
            fila.get("categoria") or "",
            fila.get("descripcion") or "",
        ], _COLOR_GASTO)
    _ajustar_columnas(hoja)


def _hoja_facturas(libro):
    hoja = libro.create_sheet("Facturas")
    filas = mod_facturas.listar_facturas(limite=5000)
    _cabecera(hoja, 1, ["ID", "Número", "Cliente", "Base (€)", "IVA %", "Total (€)", "Fecha", "Estado"])
    for i, fila in enumerate(filas, 2):
        color = _COLOR_INGRESO if fila["estado"] == "pagada" else (
            _COLOR_GASTO if fila["estado"] == "cancelada" else None
        )
        _fila_datos(hoja, i, [
            fila["id"],
            fila["numero"],
            fila.get("cliente") or "",
            float(fila["importe_base"]),
            float(fila["iva_porcentaje"]),
            float(fila["importe_total"]),
            str(fila["fecha"]),
            fila["estado"],
        ], color)
    _ajustar_columnas(hoja)


def _hoja_resumen_anual(libro, anio: int):
    hoja = libro.create_sheet("Resumen Anual")
    meses_nombre = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    _cabecera(hoja, 1, ["Mes", "Ingresos (€)", "Gastos (€)", "Balance (€)"])

    resumenes = mod_informes.resumen_anual(anio)
    total_ing = Decimal("0")
    total_gasto = Decimal("0")

    for i, r in enumerate(resumenes, 2):
        balance = r.balance
        color = _COLOR_INGRESO if balance >= 0 else _COLOR_GASTO
        _fila_datos(hoja, i, [
            meses_nombre[r.mes - 1],
            float(r.total_ingresos),
            float(r.total_gastos),
            float(balance),
        ], color)
        total_ing += r.total_ingresos
        total_gasto += r.total_gastos

    fila_total = len(resumenes) + 2
    _fila_datos(hoja, fila_total, [
        "TOTAL",
        float(total_ing),
        float(total_gasto),
        float(total_ing - total_gasto),
    ], _COLOR_TOTAL)
    for col in range(1, 5):
        hoja.cell(row=fila_total, column=col).font = Font(bold=True)

    _ajustar_columnas(hoja)


def exportar(ruta: str, anio: int):
    ruta_path = Path(ruta)
    if ruta_path.exists():
        libro = openpyxl.load_workbook(ruta)
        for nombre in ["Ingresos", "Gastos", "Facturas", "Resumen Anual"]:
            if nombre in libro.sheetnames:
                del libro[nombre]
    else:
        libro = openpyxl.Workbook()
        if "Sheet" in libro.sheetnames:
            del libro["Sheet"]

    _hoja_resumen_anual(libro, anio)
    _hoja_ingresos(libro, anio)
    _hoja_gastos(libro, anio)
    _hoja_facturas(libro)

    libro.save(ruta)
    return ruta
