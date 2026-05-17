import io

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, extract, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..database import obtener_sesion
from ..modelos.categoria import Categoria
from ..modelos.gasto import Gasto
from ..modelos.ingreso import Ingreso
from ..modelos.suscripcion import Suscripcion
from .informes import _es_mes_pago, _total_mensual_suscripciones

router = APIRouter()

_COLOR_PRIMARIO = "5C8097"
_COLOR_FILA_PAR = "F0F4F7"
_FORMATO_EURO = '#,##0.00 "€"'
_MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

_CABECERA_FILL = PatternFill(start_color=_COLOR_PRIMARIO, end_color=_COLOR_PRIMARIO, fill_type="solid")
_CABECERA_FONT = Font(color="FFFFFF", bold=True)
_FILA_PAR_FILL = PatternFill(start_color=_COLOR_FILA_PAR, end_color=_COLOR_FILA_PAR, fill_type="solid")


def _fila_cabecera(hoja, fila: int, columnas: list[str]) -> None:
    for col, nombre in enumerate(columnas, 1):
        celda = hoja.cell(row=fila, column=col, value=nombre)
        celda.font = _CABECERA_FONT
        celda.fill = _CABECERA_FILL


def _aplicar_fila_par(hoja, fila: int, n_cols: int) -> None:
    if fila % 2 == 0:
        for col in range(1, n_cols + 1):
            hoja.cell(fila, col).fill = _FILA_PAR_FILL


def _autoajustar_columnas(hoja) -> None:
    for col_cells in hoja.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        hoja.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


def _top5_categorias(filas: list) -> list[tuple[str, float]]:
    if not filas:
        return []
    resultado = [(f.nombre, float(f.total)) for f in filas[:5]]
    resto = filas[5:]
    if resto:
        resultado.append(("Otros", sum(float(f.total) for f in resto)))
    return resultado


def _agregar_grafico_categorias(
    hoja, categorias: list[tuple[str, float]], fila_inicio: int, col_grafico: str = "D"
) -> None:
    hoja.cell(fila_inicio, 1, "Top categorías de gasto").font = Font(bold=True, color=_COLOR_PRIMARIO)

    fila_cab = fila_inicio + 1
    _fila_cabecera(hoja, fila_cab, ["Categoría", "Importe (€)"])

    for i, (nombre, total) in enumerate(categorias, fila_cab + 1):
        hoja.cell(i, 1, nombre)
        hoja.cell(i, 2, total).number_format = _FORMATO_EURO
        _aplicar_fila_par(hoja, i, 2)

    fila_fin = fila_cab + len(categorias)

    grafico = BarChart()
    grafico.type = "col"
    grafico.title = "Top categorías de gasto"
    grafico.y_axis.title = "€"
    grafico.width = 15
    grafico.height = 10

    datos = Reference(hoja, min_col=2, max_col=2, min_row=fila_cab, max_row=fila_fin)
    cats = Reference(hoja, min_col=1, min_row=fila_cab + 1, max_row=fila_fin)
    grafico.add_data(datos, titles_from_data=True)
    grafico.set_categories(cats)
    hoja.add_chart(grafico, f"{col_grafico}{fila_inicio}")


async def _query_categorias_mes(db: AsyncSession, mes: int, anio: int) -> list:
    resultado = await db.execute(
        select(Categoria.nombre, func.sum(Gasto.importe).label("total"))
        .join(
            Gasto,
            and_(
                Gasto.categoria_id == Categoria.id,
                extract("month", Gasto.fecha) == mes,
                extract("year", Gasto.fecha) == anio,
            ),
        )
        .where(Categoria.tipo.in_(["gasto", "ambos"]))
        .group_by(Categoria.id, Categoria.nombre)
        .having(func.sum(Gasto.importe) > 0)
        .order_by(func.sum(Gasto.importe).desc())
    )
    return resultado.all()


async def _query_categorias_anio(db: AsyncSession, anio: int) -> list:
    resultado = await db.execute(
        select(Categoria.nombre, func.sum(Gasto.importe).label("total"))
        .join(
            Gasto,
            and_(
                Gasto.categoria_id == Categoria.id,
                extract("year", Gasto.fecha) == anio,
            ),
        )
        .where(Categoria.tipo.in_(["gasto", "ambos"]))
        .group_by(Categoria.id, Categoria.nombre)
        .having(func.sum(Gasto.importe) > 0)
        .order_by(func.sum(Gasto.importe).desc())
    )
    return resultado.all()


@router.get("/mes")
async def exportar_mes(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2000),
    db: AsyncSession = Depends(obtener_sesion),
) -> StreamingResponse:
    ingresos = (await db.execute(
        select(Ingreso)
        .where(extract("month", Ingreso.fecha) == mes, extract("year", Ingreso.fecha) == anio)
        .order_by(Ingreso.fecha)
    )).scalars().all()

    gastos = (await db.execute(
        select(Gasto)
        .where(extract("month", Gasto.fecha) == mes, extract("year", Gasto.fecha) == anio)
        .order_by(Gasto.fecha)
    )).scalars().all()

    todas_subs = (await db.execute(select(Suscripcion))).scalars().all()
    subs_mes = [s for s in todas_subs if _es_mes_pago(s, mes, anio)]

    filas_cats = await _query_categorias_mes(db, mes, anio)
    categorias = _top5_categorias(filas_cats)

    total_ingresos = sum(float(i.importe) for i in ingresos)
    total_gastos = sum(float(g.importe) for g in gastos)
    total_suscripciones = sum(float(s.importe) for s in subs_mes)
    balance = total_ingresos - total_gastos - total_suscripciones

    libro = Workbook()

    # Hoja Resumen
    hoja_resumen = libro.active
    hoja_resumen.title = "Resumen"
    _fila_cabecera(hoja_resumen, 1, ["Concepto", "Importe (€)"])
    for i, (concepto, valor) in enumerate([
        ("Ingresos", total_ingresos),
        ("Gastos", total_gastos),
        ("Suscripciones", total_suscripciones),
        ("Balance", balance),
    ], 2):
        hoja_resumen.cell(i, 1, concepto)
        hoja_resumen.cell(i, 2, valor).number_format = _FORMATO_EURO
        _aplicar_fila_par(hoja_resumen, i, 2)
    if categorias:
        _agregar_grafico_categorias(hoja_resumen, categorias, fila_inicio=7)
    _autoajustar_columnas(hoja_resumen)

    # Hoja Ingresos
    hoja_ingresos = libro.create_sheet("Ingresos")
    _fila_cabecera(hoja_ingresos, 1, ["Fecha", "Importe (€)", "Categoría", "Descripción"])
    for fila, ing in enumerate(ingresos, 2):
        hoja_ingresos.cell(fila, 1, ing.fecha).number_format = "DD/MM/YYYY"
        hoja_ingresos.cell(fila, 2, float(ing.importe)).number_format = _FORMATO_EURO
        hoja_ingresos.cell(fila, 3, ing.categoria.nombre if ing.categoria else "")
        hoja_ingresos.cell(fila, 4, ing.descripcion or "")
        _aplicar_fila_par(hoja_ingresos, fila, 4)
    _autoajustar_columnas(hoja_ingresos)

    # Hoja Gastos
    hoja_gastos = libro.create_sheet("Gastos")
    _fila_cabecera(hoja_gastos, 1, ["Fecha", "Importe (€)", "Categoría", "Descripción"])
    for fila, gasto in enumerate(gastos, 2):
        hoja_gastos.cell(fila, 1, gasto.fecha).number_format = "DD/MM/YYYY"
        hoja_gastos.cell(fila, 2, float(gasto.importe)).number_format = _FORMATO_EURO
        hoja_gastos.cell(fila, 3, gasto.categoria.nombre if gasto.categoria else "")
        hoja_gastos.cell(fila, 4, gasto.descripcion or "")
        _aplicar_fila_par(hoja_gastos, fila, 4)
    _autoajustar_columnas(hoja_gastos)

    # Hoja Suscripciones
    hoja_subs = libro.create_sheet("Suscripciones")
    _fila_cabecera(hoja_subs, 1, ["Nombre", "Importe (€)", "Frecuencia", "Día de cobro"])
    for fila, sub in enumerate(subs_mes, 2):
        hoja_subs.cell(fila, 1, sub.nombre)
        hoja_subs.cell(fila, 2, float(sub.importe)).number_format = _FORMATO_EURO
        hoja_subs.cell(fila, 3, sub.frecuencia)
        hoja_subs.cell(fila, 4, sub.dia_cobro or "")
        _aplicar_fila_par(hoja_subs, fila, 4)
    _autoajustar_columnas(hoja_subs)

    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)

    nombre_fichero = f"contabilidad_{anio}_{mes:02d}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_fichero}"'},
    )


@router.get("/anual/{anio}")
async def exportar_anual(
    anio: int,
    db: AsyncSession = Depends(obtener_sesion),
) -> StreamingResponse:
    filas_ingresos = (await db.execute(
        select(
            extract("month", Ingreso.fecha).label("mes"),
            func.sum(Ingreso.importe).label("total"),
        )
        .where(extract("year", Ingreso.fecha) == anio)
        .group_by(extract("month", Ingreso.fecha))
    )).all()

    filas_gastos = (await db.execute(
        select(
            extract("month", Gasto.fecha).label("mes"),
            func.sum(Gasto.importe).label("total"),
        )
        .where(extract("year", Gasto.fecha) == anio)
        .group_by(extract("month", Gasto.fecha))
    )).all()

    ingresos_por_mes = {int(f.mes): float(f.total) for f in filas_ingresos}
    gastos_por_mes = {int(f.mes): float(f.total) for f in filas_gastos}

    meses_data = []
    for m in range(1, 13):
        sus = float(await _total_mensual_suscripciones(db, m, anio))
        ing = ingresos_por_mes.get(m, 0.0)
        gas = gastos_por_mes.get(m, 0.0)
        meses_data.append((m, ing, gas, sus, ing - gas - sus))

    total_ing = sum(r[1] for r in meses_data)
    total_gas = sum(r[2] for r in meses_data)
    total_sus = sum(r[3] for r in meses_data)
    total_bal = total_ing - total_gas - total_sus

    filas_cats = await _query_categorias_anio(db, anio)
    categorias = _top5_categorias(filas_cats)

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Resumen"

    _fila_cabecera(hoja, 1, ["Mes", "Ingresos (€)", "Gastos (€)", "Suscripciones (€)", "Balance (€)"])
    for fila, (m, ing, gas, sus, bal) in enumerate(meses_data, 2):
        hoja.cell(fila, 1, _MESES[m - 1])
        for col, valor in enumerate([ing, gas, sus, bal], 2):
            hoja.cell(fila, col, valor).number_format = _FORMATO_EURO
        _aplicar_fila_par(hoja, fila, 5)

    fila_total = 14
    hoja.cell(fila_total, 1, "TOTAL").font = Font(bold=True)
    for col, valor in enumerate([total_ing, total_gas, total_sus, total_bal], 2):
        celda = hoja.cell(fila_total, col, valor)
        celda.number_format = _FORMATO_EURO
        celda.font = Font(bold=True)
    _aplicar_fila_par(hoja, fila_total, 5)

    if categorias:
        _agregar_grafico_categorias(hoja, categorias, fila_inicio=16, col_grafico="G")
    _autoajustar_columnas(hoja)

    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="contabilidad_{anio}.xlsx"'},
    )
